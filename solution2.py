import openpyxl as xl


def xl_to_dict(sheet):
    dictionary = {}
    for i in range(2, sheet.max_row + 1):
        english_word = sheet.cell(i, 1).value
        for j in range(2, sheet.max_column + 1):
            if sheet.cell(i, j).value != None:
                punjabi_word = sheet.cell(i, j).value.replace(" ","")
                dictionary[punjabi_word] = english_word
    return dictionary


def search_in_dict(dictionary, word):
    word = word.replace(' ', '')
    meaning = dictionary.get(word)
    if meaning == None:
        return 'Word/Sentence not found in dictionary'
    else:
        return meaning

##########################################################################################

xl_sheet = xl.load_workbook('Test1.xlsx')
sheet = xl_sheet['Sheet1']

mydict = xl_to_dict(sheet)

print(search_in_dict(mydict, ' ਜੌਹਰੀਆਂ ਦਾ ਔਜ਼ਾਰ ਜਿਸ ਨਾਲ ਜਵਾਹਰਾਂ ਨੂੰ ਕਟਿਆ ਜਾਂਦਾ   ਹੈ'))
print(search_in_dict(mydict, 'abc'))    # 'abc' is not present in our excel and dictionary